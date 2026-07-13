#!/usr/bin/env python3
"""export.py — Spreadsheet output stage of the Deep Research Agent.

Reads the most recent ranking snapshot from `rankings_history` (written by
analyze.py) and writes it out as both a plain CSV and a styled `.xlsx`
workbook. Re-running the pipeline overwrites these files with the latest
snapshot, while the full history stays queryable in SQLite.

Usage:
    python export.py
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src import db
from src.config import config

RISK_FILL = {
    "Low": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "Medium": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "High": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
}
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)

COLUMNS = [
    ("overall_rank", "Overall Rank"),
    ("position_rank", "Pos Rank"),
    ("name", "Player"),
    ("position", "Position"),
    ("team", "Team"),
    ("projected_points", "Projected Points"),
    ("sos_adjustment", "SOS Adj %"),
    ("risk_tier", "Risk Tier"),
    ("tag", "Tag"),
    ("notes", "Notes"),
]


def _latest_rankings_dataframe() -> tuple[pd.DataFrame, str | None]:
    with db.connect(config.db_path) as conn:
        run_timestamp = db.latest_run_timestamp(conn)
        if not run_timestamp:
            return pd.DataFrame(columns=[c[1] for c in COLUMNS]), None
        rows = db.fetch_rankings_for_run(conn, run_timestamp)

    records = []
    for row in rows:
        d = dict(row)
        d["sos_adjustment"] = round(d["sos_adjustment"] * 100, 1)
        records.append({label: d[key] for key, label in COLUMNS})

    df = pd.DataFrame(records, columns=[c[1] for c in COLUMNS])
    return df, run_timestamp


def export_csv(df: pd.DataFrame) -> Path:
    df.to_csv(config.csv_output, index=False)
    return config.csv_output


def export_xlsx(df: pd.DataFrame, run_timestamp: str | None) -> Path:
    config.xlsx_output.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(config.xlsx_output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Rankings")
        ws = writer.sheets["Rankings"]

        for col_idx, (_key, label) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")

        risk_col_idx = [i for i, (key, _l) in enumerate(COLUMNS, start=1) if key == "risk_tier"][0]
        for row_idx in range(2, ws.max_row + 1):
            risk_value = ws.cell(row=row_idx, column=risk_col_idx).value
            fill = RISK_FILL.get(risk_value)
            if fill:
                for col_idx in range(1, len(COLUMNS) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = fill

        for col_idx, (_key, label) in enumerate(COLUMNS, start=1):
            max_len = max([len(label)] + [len(str(v)) for v in df[label].astype(str)]) if len(df) else len(label)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

        ws.freeze_panes = "A2"
        if run_timestamp:
            meta_row = ws.max_row + 2
            ws.cell(row=meta_row, column=1, value=f"Generated: {run_timestamp}")

    return config.xlsx_output


def run_export() -> dict:
    df, run_timestamp = _latest_rankings_dataframe()
    csv_path = export_csv(df)
    xlsx_path = export_xlsx(df, run_timestamp)
    return {
        "run_timestamp": run_timestamp,
        "rows": len(df),
        "csv_path": str(csv_path),
        "xlsx_path": str(xlsx_path),
    }


def main() -> None:
    summary = run_export()
    print(json.dumps(summary, indent=2))


if __name__ == "__main__":
    main()
